from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # 기본이름으로 sheet 생성
ws.title = "MySheet" # sheet 이름 변경

ws.sheet_properties.tabColor = "ff66ff"
#색은 https://www.w3schools.com/colors/colors_picker.asp?colorhex=adb568 를 참고하거나
#구글에서 RGB를 치면된다


ws1 = wb.create_sheet("YourSheet") #주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("NewSheet", 2) #2번째에 sheet 를 생성

new_ws = wb["NewSheet"] #Dict 형태로 Sheet 에 접근 가능하다

print(wb.sheetnames) #모든 sheet 이름 확인
new_ws["A1"] = "Test" # A1 에 값을 입력
target = wb.copy_worksheet(new_ws) #worksheet 복사
target.title = "Copied Sheet" # sheet 이름 설정

wb.save("sample .xlsx")
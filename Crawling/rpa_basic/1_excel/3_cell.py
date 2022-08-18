from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3
ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) # A1셀의 정보 출력 (<Cell 'NadoSheet'.A1>)
print(ws["A1"].value) # A1셀의 "값"을 출력 (1)
print(ws["A10"].value) # 값이 없을 땐 None 을 출력

print(ws.cell(column=1, row=1).value) #print(ws["A1"])
print(ws.cell(column=2, row=1).value) #print(ws["B1"])

ws.cell(column=3, row=1, value=10) #ws["C1"]=10

#반복문을 통해 값 넣기

from random import *

index = 1
for x in range(1, 11):
    for y in range(1, 11):
        #ws.cell(row=x, column=y, value=randint(0,100)) #0~99 사이 숫자 넣음
        ws.cell(row=x, column=y, value=index)
        index +=1


wb.save("sample.xlsx")
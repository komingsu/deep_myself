from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

#1줄씩 데이터 넣기
ws.append(["번호", "영어" , "수학"])
for i in range(1,11):
    ws.append([i, randint(0,100), randint(0,100)])

col_B = ws["B"] # B 칼럼을 다 가져옴

for cell in col_B:
    print(cell.value)

col_range = ws["B:C"] #B~C 칼럼 가져오기
for cols in col_range:
    for cell in cols:
        print(cell.value)


row_title = ws[1]
for cell in row_title:
    print(cell.value)

row_range = ws[2:6] # 6을 포함해서 가져온다
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
    print()

from openpyxl.utils.cell import coordinate_from_string # 현재 sheet 좌표정보 가져오기


row_range = ws[2:ws.max_row]
for rows in row_range:
    for cell in rows:
        #print(cell.value, end=" ")
        #print(cell.coordinate, end=" ")
        xy = coordinate_from_string(cell.coordinate)
        print(xy, end=" ")
    print()

print(tuple(ws.rows)) #A~C1, A~C2, A~C3 ...
print(tuple(ws.columns)) #A1~10, B1~10 ,C1~10 ...

# for row in tuple(ws.rows):
#     print(row[2])

for row in ws.iter_rows(): #전체 row
    print(row[0].value) #A열가져옴

# for row in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3): #가져오는 행,열 의 최대, 최소를 정할수 있음
#     print(row[0])

# for col in ws.iter_cols(): #가져오는 행,열 의 최대, 최소를 정할수 있음
#     print(col)


wb.save("sample.xlsx")
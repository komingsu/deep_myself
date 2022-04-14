import requests
from bs4 import BeautifulSoup
import re
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook #파일 불러오기
from openpyxl import Workbook
import csv
from os import write

# ## 읽어온자료 쓸 준비하기
# filename = "커피숍_분석데이터.csv"
# f = open(filename, "w", encoding="utf-8-sig", newline="") #newline 을 공백으로 해줘야 한줄띄기가 없어진다.
# writer = csv.writer(f)

# title = "N	종목명	현재가	전일비	등락률	액면가	시가총액	상장주식수	외국인비율	거래량	PER	ROE".split("\t")
# #탭으로 구분한 title이 된다.
# writer.writerow(title)


wb = load_workbook("coffeeshop_data.xlsx") #파일을 불러옴
ws = wb.active #활성화된 sheet

col_F = ws["F"] # B 칼럼을 다 가져옴

filename = "커피숍_분석데이터자료.csv"
f = open(filename, "w", encoding="utf-8-sig", newline="") #newline 을 공백으로 해줘야 한줄띄기가 없어진다.
#여기서 한글이 깨질경우에는 encoding="utf-8-sig" 로 해주면 된다
writer = csv.writer(f)
# title = "N	종목명	현재가	전일비	등락률	액면가	시가총액	상장주식수	외국인비율	거래량	PER	ROE".split("\t")
# #탭으로 구분한 title이 된다.
# writer.writerow(title)

headers = {"User-Agent" : "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.61 Safari/537.36"}

# 첫데이터 가져오기

# 브라우저 켜기
browser = webdriver.Chrome()
browser.maximize_window()
url = "https://sg.sbiz.or.kr/godo/index.sg"
browser.get(url)

# 로그인 하기
browser.find_element_by_xpath("//*[@id='help_guide']/div/div[2]/div[2]/label").click()
browser.find_element_by_xpath("//*[@id='help_guide']/div/div[2]/a/span").click()
time.sleep(1)
browser.find_element_by_xpath("//*[@id='currentState']/a/h4").click()
time.sleep(1)
browser.find_element_by_xpath("//*[@id='toLink']/a/h4").click()
time.sleep(1)
browser.find_element_by_xpath("//*[@id='wrap']/div/div/div[3]/a[2]/span").click()
time.sleep(1)
browser.find_element_by_id("id").send_keys("rh9872")
time.sleep(2)
browser.find_element_by_id("pass").send_keys("alstnfpdh2!@")
browser.find_element_by_xpath("/html/body/div/div[3]/form/div/button").click()
time.sleep(3)

#상세분석 들어가기
browser.find_element_by_xpath("//*[@id='currentState']/a/h4").click()
time.sleep(1)
browser.find_element_by_xpath("//*[@id='toLink']/a/h4").click()
time.sleep(1)

# 옵션 설정

browser.find_element_by_xpath("//*[@id='upjongSelection']").click()
time.sleep(1)
browser.find_element_by_id("searchWord").send_keys("카페")
time.sleep(1)
browser.find_element_by_xpath("//*[@id='select_location_1']").click()
time.sleep(1)
browser.find_element_by_xpath("//*[@id='checkTypeConfirm']/span").click()
time.sleep(3)

# 주소 기입
browser.find_element_by_xpath("//*[@id='searchAddress']").send_keys(add)
time.sleep(1)
browser.find_element_by_xpath("//*[@id='layerPopAddressMove']").click()

# 분석 설정
time.sleep(1)
browser.find_element_by_xpath("//*[@id='map']/div[1]/div/div[6]/div[2]/div/ul/li[1]/label").click()
time.sleep(1)
browser.find_element_by_xpath("//*[@id='map']/div[1]/div/div[6]/div[2]/div/ul/li[1]/div/ul/li[2]/label").click()
time.sleep(1)
browser.find_element_by_xpath("//*[@id='auto_circle_1']").click()
time.sleep(1)
browser.find_element_by_xpath("//*[@id='auto_circle']/div/div[3]/a[2]/span").click()
time.sleep(1)
browser.find_element_by_xpath("//*[@id='map']/div[1]/div/div[6]/div[3]/img").click()
time.sleep(5)



# for add in col_F[1:5] :
#     print(add.value)
# from openpyxl import load_workbook #파일 불러오기
# from openpyxl import Workbook




# wb = load_workbook("coffeeshop_data.xlsx") #파일을 불러옴
# ws = wb.active #활성화된 sheet

# col_F = ws["F"]

# for add in col_F[1:5] :
#     print(add.value)

# import csv
# from os import write
# import requests
# from bs4 import BeautifulSoup

# filename = "연습.csv"
# f = open(filename, "w", encoding="utf-8-sig", newline="") #newline 을 공백으로 해줘야 한줄띄기가 없어진다.
# #여기서 한글이 깨질경우에는 encoding="utf-8-sig" 로 해주면 된다
# writer = csv.writer(f)

# title = "M_08_sales M_08_sales_count    O_08_sales  O_08_sales_count    M_07_sales M_07_sales_count    O_07_sales  O_07_sales_count M_06_sales M_06_sales_count    O_06_sales  O_06_sales_count M_05_sales M_05_sales_count    O_05_sales  O_05_sales_count Main_buyer_man  Main_buyer_age  Main_buyer_weekday  Floating_POP    Floating_POP_man    Floating_POP_weekday    Floating_POP_age    Residential_POP Residential_POP_man Residential_POP_age Office_POP  Office_POP_sex  Office_POP_age  Growth  Stability   Sales   Buying  Attention   Office_spend    Residential_spend   Key_facility".split("\t")
# writer.writerow(title)

import csv
from os import write
import requests
import re
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook #파일 불러오기
from openpyxl import Workbook
from bs4 import BeautifulSoup


f= open('sample2.csv', 'w', encoding='utf-8-sig', newline='') #CP949, MS949, EUC-KR
wr = csv.writer(f)

browser = webdriver.Chrome('chromedriver')
browser.maximize_window()
url = "https://www.naver.com/"
browser.get(url)
time.sleep(2)

soup = BeautifulSoup(browser.page_source, "lxml")
#movies = soup.find_all("div", attrs= {"class" :  "Vpfmgd"})
movies = WebDriverWait(browser,5).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR,"#NM_FAVORITE > div.group_nav > ul.list_nav.NM_FAVORITE_LIST > li:nth-child(1) > a"))).get_text()
movies = soup.select_one("#NM_FAVORITE > div.group_nav > ul.list_nav.NM_FAVORITE_LIST > li:nth-child(1) > a").get_text()
print(movies)
wr.writerow([movies])
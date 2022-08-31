import csv
from os import write
from bs4 import BeautifulSoup
import re
import time
from selenium import webdriver
#from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
#from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook #파일 불러오기
from openpyxl import Workbook
import urllib.request as req

### csv로 데이터를 받아올 준비 ######

filename = "커피숍분석자료.csv"
f = open(filename, "w", encoding="utf-8-sig", newline="")
writer = csv.writer(f)
title = ["영업장수","분석8월매출액","분석8월건수","유사8월매출액","유사8월건수","분석7월매출액","분석7월건수","유사7월매출액","유사7월건수","분석6월매출액","분석6월건수","유사6월매출액","유사6월건수","분석5월매출액","분석5월건수","유사5월매출액","유사5월건수","주요구매성비","주요구매나이","주중구매비율","유동인구","유동인구_남성","유동인구_주중","유동인구_나이","거주인구","거주인구_남성","거주인구_나이","직장인구","직장인구_남성","직장인구_나이","성장성지수","안정성지수","영업력지수","구매력","집객력","주요시설수","주소"]
writer.writerow(title)

####################################

### col_F 에 주소 값 삽입 ###########

wb = load_workbook("coffeeshop_data.xlsx") #파일을 불러옴
ws = wb.active
col_F = ws["F"]

####################################

headers = {"User-Agent" : "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.61 Safari/537.36"}

for inter in range(0,10) :

    ### 브라우저 켜기 ##################
    browser = webdriver.Chrome()
    browser.maximize_window()
    url = "https://sg.sbiz.or.kr/godo/index.sg"
    browser.get(url)
    browser.implicitly_wait(50)
    time.sleep(2)

    ### 로그인 하기
    browser.find_element_by_xpath("/html/body/div/div[2]/div[20]/div[3]/label[1]").click()
    time.sleep(1)
    browser.find_element_by_xpath("//*[@id='help_guide']/div/div[2]/div[2]/label").click()
    time.sleep(1)
    browser.find_element_by_xpath("//*[@id='help_guide']/div/div[2]/a/span").click()
    time.sleep(1)
    browser.find_element_by_xpath("//*[@id='currentState']/a/h4").click()
    time.sleep(1)
    browser.find_element_by_xpath("//*[@id='toLink']/a/h4").click()
    time.sleep(1)
    browser.find_element_by_xpath("//*[@id='wrap']/div/div/div[3]/a[2]/span").click()
    time.sleep(1)
    browser.find_element_by_id("id").send_keys("rh9872")
    time.sleep(1)
    browser.find_element_by_id("pass").send_keys("alstnfpdh2!@")
    time.sleep(1)
    browser.find_element_by_xpath("/html/body/div/div[3]/form/div/button").click()
    time.sleep(1)


    ### 상세분석 들어가기
    browser.find_element_by_xpath("//*[@id='currentState']/a/h4").click()
    time.sleep(1.5)
    browser.find_element_by_xpath("//*[@id='toLink']/a/h4").click()
    time.sleep(3)


    ### 옵션 설정

    browser.find_element_by_xpath("//*[@id='upjongSelection']").click()
    time.sleep(1)
    browser.find_element_by_id("searchWord").send_keys("카페")
    time.sleep(1)
    browser.find_element_by_xpath("//*[@id='select_location_1']").click()
    time.sleep(1)
    browser.find_element_by_xpath("//*[@id='checkTypeConfirm']/span").click()
    time.sleep(1)

    ##################################

    ### 주소 기입 ####################

    #1 -> 2 부터 체크시작
    
    for add in col_F[inter*90 + 4813: inter*90 + 4725 ] :
        
        time.sleep(2)
        address = add.value
        
        # 주소 입력
        browser.find_element_by_xpath("/html/body/div/div[2]/div[1]/div[2]/div/input").send_keys(address)
        time.sleep(1)
        # 확인
        browser.find_element_by_xpath("/html/body/div/div[2]/div[1]/div[2]/div/a").click()
        time.sleep(1)
        # 상권분석
        browser.find_element_by_xpath("/html/body/div/div[2]/div[10]/div[1]/div/div[6]/div[2]/div/ul/li[1]/label").click()
        time.sleep(1)
        # 반경
        browser.find_element_by_xpath("/html/body/div/div[2]/div[10]/div[1]/div/div[6]/div[2]/div/ul/li[1]/div/ul/li[2]/label/span").click()
        time.sleep(1)
        # 500m
        browser.find_element_by_xpath("/html/body/div/div[2]/div[8]/div/div[2]/ul/li[5]/label").click()
        time.sleep(1)
        # 확인
        browser.find_element_by_xpath("/html/body/div/div[2]/div[8]/div/div[3]/a[2]/span").click()
        time.sleep(2)
        # 분석
        browser.find_element_by_xpath("/html/body/div/div[2]/div[10]/div[1]/div/div[6]/div[3]/img").click()
        time.sleep(45)

        ### 분석보고서 속성값들 가져오기
        soup = BeautifulSoup(browser.page_source, "html.parser")
        Num_Store = soup.select_one("#page2 > div.table-wrap > table > tbody > tr:nth-child(1) > td:nth-child(7) > div > strong").get_text()
        ### 매출분석 탭

        M_08_sales = soup.select_one("#page3 > div:nth-child(12) > table > tbody > tr:nth-child(1) > td:nth-child(8) > div > strong").get_text()
        M_08_sales_count = soup.select_one("#page3 > div:nth-child(12) > table > tbody > tr:nth-child(2) > td:nth-child(7) > div > strong").get_text()
        O_08_sales = soup.select_one("#page3 > div:nth-child(12) > table > tbody > tr:nth-child(3) > td:nth-child(8) > div > strong").get_text()
        O_08_sales_count = soup.select_one("#page3 > div:nth-child(12) > table > tbody > tr:nth-child(4) > td:nth-child(7) > div > strong").get_text()
        M_07_sales = soup.select_one("#page3 > div:nth-child(12) > table > tbody > tr:nth-child(1) > td:nth-child(7) > div > strong").get_text()
        M_07_sales_count = soup.select_one("#page3 > div:nth-child(12) > table > tbody > tr:nth-child(2) > td:nth-child(6) > div > strong").get_text()
        O_07_sales = soup.select_one("#page3 > div:nth-child(12) > table > tbody > tr:nth-child(3) > td:nth-child(7) > div > strong").get_text()
        O_07_sales_count = soup.select_one("#page3 > div:nth-child(12) > table > tbody > tr:nth-child(4) > td:nth-child(6) > div > strong").get_text()
        M_06_sales = soup.select_one("#page3 > div:nth-child(12) > table > tbody > tr:nth-child(1) > td:nth-child(6) > div > strong").get_text()
        M_06_sales_count = soup.select_one("#page3 > div:nth-child(12) > table > tbody > tr:nth-child(2) > td:nth-child(5) > div > strong").get_text()
        O_06_sales = soup.select_one("#page3 > div:nth-child(12) > table > tbody > tr:nth-child(3) > td:nth-child(6) > div > strong").get_text()
        O_06_sales_count = soup.select_one("#page3 > div:nth-child(12) > table > tbody > tr:nth-child(4) > td:nth-child(5) > div > strong").get_text()
        M_05_sales = soup.select_one("#page3 > div:nth-child(12) > table > tbody > tr:nth-child(1) > td:nth-child(5) > div > strong").get_text()
        M_05_sales_count = soup.select_one("#page3 > div:nth-child(12) > table > tbody > tr:nth-child(2) > td:nth-child(4) > div > strong").get_text()
        O_05_sales = soup.select_one("#page3 > div:nth-child(12) > table > tbody > tr:nth-child(3) > td:nth-child(5) > div > strong").get_text()
        O_05_sales_count = soup.select_one("#page3 > div:nth-child(12) > table > tbody > tr:nth-child(4) > td:nth-child(4) > div > strong").get_text()
        Main_buyer_man = soup.select_one("#page3 > div:nth-child(28) > table > tbody > tr:nth-child(2) > td:nth-child(2)").get_text()
        Main_buyer_age = soup.select_one("#page3 > div:nth-child(29) > div.midd > ul > li:nth-child(2) > span:nth-child(1)").get_text()
        Main_buyer_weekday = soup.select_one("#page3 > div:nth-child(19) > table > tbody > tr:nth-child(2) > td:nth-child(2)").get_text()

        ### 인구분석 탭
        #browser.find_element_by_xpath("/html/body/div[2]/div/div[2]/div[2]/div[1]/div/ul/li[3]/a/span").click()
        Floating_POP = soup.select_one("#page4 > div:nth-child(9) > table > tbody > tr:nth-child(1) > td:nth-child(3)").get_text()
        Floating_POP_man = soup.select_one("#page4 > div:nth-child(9) > table > tbody > tr:nth-child(2) > td:nth-child(2)").get_text()
        Floating_POP_weekday = soup.select_one("#page4 > div:nth-child(12) > table > tbody > tr:nth-child(2) > td:nth-child(2)").get_text()
        Floating_POP_age = soup.select_one("#page4 > div:nth-child(15) > div.midd > ul > li:nth-child(1) > span:nth-child(4)").get_text()
        #Residential_POP Residential_POP_man Residential_POP_age Office_POP  Office_POP_sex  Office_POP_age
        Residential_POP = soup.select_one("#page4 > div:nth-child(19) > table > tbody > tr:nth-child(1) > td:nth-child(3)").get_text()
        Residential_POP_man = soup.select_one("#page4 > div:nth-child(19) > table > tbody > tr:nth-child(2) > td:nth-child(3)").get_text()
        Residential_POP_age = soup.select_one("#page4 > div:nth-child(20) > div.midd > ul > li > span:nth-child(5)").get_text()
        Office_POP = soup.select_one("#page4 > div:nth-child(25) > div.midd > ul > li > span:nth-child(2)").get_text()
        Office_POP_man = soup.select_one("#page4 > div:nth-child(24) > table > tbody > tr:nth-child(2) > td:nth-child(3)").get_text()
        Office_POP_age = soup.select_one("#page4 > div:nth-child(25) > div.midd > ul > li > span:nth-child(5)").get_text()
        #Growth  Stability   Sales   Buying  Attention   Office_spend    Residential_spend   Key_facility
        ### 상권평가 탭
        #browser.find_element_by_xpath("/html/body/div[2]/div/div[2]/div[2]/div[1]/div/ul/li[6]/a/span").click()
        Growth = soup.select_one("#page1 > div:nth-child(4) > table > tbody > tr > td:nth-child(1)").get_text()
        Stability = soup.select_one("#page1 > div:nth-child(4) > table > tbody > tr > td:nth-child(2)").get_text()
        Sales = soup.select_one("#page1 > div:nth-child(4) > table > tbody > tr > td:nth-child(3)").get_text()
        Buying = soup.select_one("#page1 > div:nth-child(4) > table > tbody > tr > td:nth-child(4)").get_text()
        Attention = soup.select_one("#page1 > div:nth-child(4) > table > tbody > tr > td:nth-child(5)").get_text()
        ### 지역분석
        #browser.find_element_by_xpath("/html/body/div[2]/div/div[2]/div[2]/div[1]/div/ul/li[5]/a/span").click()
        Key_facility = soup.select_one("#page6 > div:nth-child(6) > table > tbody > tr > td:nth-child(3)").get_text()
        time.sleep(3)
        writer.writerow([Num_Store,M_08_sales,M_08_sales_count,O_08_sales,O_08_sales_count,M_07_sales,M_07_sales_count,O_07_sales,O_07_sales_count,M_06_sales,M_06_sales_count,O_06_sales,O_06_sales_count,M_05_sales,M_05_sales_count,O_05_sales,O_05_sales_count,Main_buyer_man,Main_buyer_weekday,Main_buyer_age,Floating_POP,Floating_POP_man,Floating_POP_weekday,Floating_POP_age,Residential_POP,Residential_POP_man,Residential_POP_age,Office_POP,Office_POP_man,Office_POP_age,Growth,Stability,Sales,Buying,Attention,Key_facility,address])
        ### 상세보고서 닫기
        browser.find_element_by_xpath("/html/body/div[2]/div/a/span").click()
        time.sleep(2)
        ### 주소자리 비우기
        browser.find_element_by_xpath("//*[@id='searchAddress']").clear()
        time.sleep(2)
        
    
    browser.quit()
    time.sleep(2)



        
    #######################################################

